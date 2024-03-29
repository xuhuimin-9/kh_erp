import web, datetime
import time
from decimal import Decimal
from datetime import datetime
import openpyxl
import pymysql

import xlwt


db = web.database(dbn="mysql", db="kh_erp", host="localhost" , user="root" , password="123456")

# 根据表名查询表数据
def select_table_from_sql(table_name):
    return db.select(table_name)

# 插入库存表 目前不用
def insert_info_to_material_io_storage_info(category_id, material_id, name, unit, count, price, tax,tax_price):
    table_name = "material_io_storage_info";
    db.insert(
        table_name, category_id=category_id, material_id=material_id, name=name, unit=unit, count=count, price=price,
        tax_rate=tax, tax_price=tax_price
    )

# 更新商品表数量 目前不用
def update_count_to_kh_material(category_id,material_id,unit,count):
    table_name = "kh_material";
    current_count = db.select(
        table_name, where="category_id=$category_id AND material_id=$material_id AND unit = $unit", vars=locals()
    )[0]['count']
    new_count = int(current_count) + int(count);
    db.update(table_name, where="category_id=$category_id AND material_id=$material_id AND unit = $unit",
              vars=locals(), count=new_count)

# 商品入库
def material_in_storage(current_material):

    category_id = current_material['categoryId']
    category_name = current_material['categoryName']
    material_id = current_material['materialId']
    name = current_material['materialName']
    unit = current_material['materialUnit']
    count = current_material['materialCount']
    price = current_material['materialPrice']
    tax = current_material['materialTax']
    tax_price = current_material['materialTaxPrice']
    invoice = current_material['invoice']
    storage = current_material['storageName']
    supplier = current_material['supplier']

    '''kh_material不将累计数量存在该表中
    # 更新商品库存数量
    table_name = "kh_material";
    # 商品表中商品信息
    material = db.select(
        table_name, where="category_id=$category_id AND material_id=$material_id AND unit = $unit", vars=locals()
    )
    if (not material):
        return 0;  # 商品表中无该条信息
    # 该条商品信息的当前数量
    current_count = material[0]['count'];
    new_count = int(current_count) + int(count);
    status = db.update(table_name, where="category_id=$category_id AND material_id=$material_id AND unit = $unit",
                          vars=locals(), count=new_count)
    if (not status):
        return -1;  # 插入kh_material失败'''

    '''     1.存入库存信息表       '''
    table_name = "material_io_storage_info";
    status = db.insert(
        table_name, category_id=category_id, category_name=category_name, material_id=material_id, name=name, unit=unit,
        count=count, price=price,tax_rate=tax, tax_price=tax_price, invoice_type = invoice,storage_name=storage,supplier=supplier)
    if(not status):
        return -2 # 插入material_io_storage_info失败

    '''     2.存入入库日志表       '''
    table_name = "material_in_storage_log";
    status = db.insert(
        table_name, category_id=category_id, category_name=category_name, material_id=material_id, name=name, unit=unit,
        count=count, price=price,tax_rate=tax, tax_price=tax_price, invoice_type = invoice,storage_name=storage,supplier=supplier)
    if (not status):
        return -3  # 插入material_in_storage_log失败

    '''     3.更改商品表状态       '''
    table_name = "kh_material";
    material = db.select(
        table_name, where="category_id=$category_id AND material_id=$material_id AND unit = $unit", vars=locals()
    )
    # 如果存在该条商品，则获取该条商品的状态【0：库存中未入库；1：已入库】
    if material :
        material_status = material[0]['status'];
    else:
        return -4   # kh_material表中未找到该商品信息
    if (not material_status ):
        status = db.update(table_name, where="category_id=$category_id AND material_id=$material_id AND unit = $unit",
                           vars=locals(), status=1)
        if (not status) :
            return -5   # 更新kh_material表中状态异常

    return 1


# 商品出库
def material_out_storage(current_material):

    # 该id为material_io_storage_info表的中唯一id，非商品id
    id = current_material['id']
    # 出库数量
    outCount = current_material['outCount']
    # 所用项目
    project = current_material['project']
    # 申领仓库
    apply_storage = current_material['applyStorage']

    '''     1.更新库存信息【数量、含税总价】        '''
    table_name = "material_io_storage_info";
    value = db.select(table_name, where="id=$id", vars=locals())
    if (not value):
        return -1  # 商品表中无该条信息
    # webpy的sql结果只能遍历一次，转为list可重复使用
    material = list(value)

    '''     更改该条入库信息状态，不允许删除        '''
    # 该条商品入库信息的状态（入库时间一致且商品名称一致）
    current_status = list(db.select("material_in_storage_log",
                                    where="create_time=$material[0]['create_time'] AND name=$material[0]['name'] AND count=$material[0]['count']",
                                    vars=locals()))
    if len(current_status) == 1 and current_status[0]['status'] == 0:
        db.update("material_in_storage_log",
                  where="create_time=$material[0]['create_time'] AND name=$material[0]['name'] AND count=$material[0]['count']",
                  vars=locals(), status=1)  # 更新状态标志位

    # 该条商品信息的当前库存
    current_count = material[0]['count']
    new_count = int(current_count) - int(outCount)  # 当前数量减去出库数量

    # 该条商品的入库含税总价
    current_tax_price = material[0]['tax_price'];
    new_tax_price = current_tax_price / current_count * new_count;  # 获得原先含税单价并乘上现有个数

    status = db.update(table_name, where="id=$id",vars=locals(), count=new_count,tax_price=new_tax_price)
    if (not status):
        return -2;  # kh_material更新失败


    '''     2.存入出库日志表       '''
    # 该条入库记录的入库时间 当作批次号
    batch=material[0]['create_time']

    category_id = material[0]['category_id']
    category_name = material[0]['category_name']
    material_id = material[0]['material_id']
    material_name = material[0]['name']
    unit = material[0]['unit']
    price = material[0]['price']
    tax = material[0]['tax_rate']
    invoice = material[0]['invoice_type']
    storage = material[0]['storage_name']
    supplier = material[0]['supplier']

    value = float(Decimal(str(material[0]['tax_price'] / material[0]['count'])))  #单个含税价
    #value = float(Decimal(str(1 + tax)) * Decimal(str(price)))  #单个含税价
    tax_price = float(Decimal(str(value * int(outCount))))      #含税总价

    # 专票商品出库时的金额采用入库的未税单价金额
    if (material[0]['invoice_type'] == "专票"):
        total_price = float(Decimal(str(price * int(outCount))))
    else:
        price = value
        total_price = tax_price

    table_name = "material_out_storage_log"
    # 如申领仓库与实际仓库一致，则正常出库
    if(storage==apply_storage):
        statusValue="NL"
    else:
        statusValue="BORROW"
    status = db.insert(table_name,
                       category_id=category_id, category_name=category_name, material_id=material_id,
                       name=material_name,unit=unit, count=outCount, price=price, invoice_type=invoice, tax_rate=tax,
                       total_price=total_price,tax_price=tax_price, project=project, storage_name=storage,
                       supplier=supplier,apply_storage=apply_storage, material_batch=batch,status=statusValue
                       )
    if (not status):
        return -4  # 插入material_in_storage_log失败

    '''     3.更新到项目表中 material_id字段值 【含义：商品id 商品名 商品数量 出库时间】        '''

    # 包含当前日期和时间的datetime对象
    #now = datetime.now()
    # dd/mm/YY H:M:S
    #dt_string = now.strftime("%Y/%m/%d-%H:%M:%S")

    #table_name = "kh_project"
    #value = db.select(table_name, where="name=$project", vars=locals())[0]['material_id']
    #value = value + str(material_id) + " " + material_name + " " + outCount + " " + dt_string +";"

    #status = db.update(table_name, where="name=$project",vars=locals(), material_id=value)
    #if (not status):
    #    return -5  # 插入kh_project失败

    # [项目名称 商品id 商品名称 出库数量 出库单价(区分专普) 出库税率 出库未税总价 出库含税总价 仓库名称]
    table_name = "kh_project_material"
    status = db.insert(table_name, project_name=project, material_id=material_id, material_name=material_name,
                       outcount=outCount, price=price, invoice = invoice, total_price=total_price,tax_price=tax_price,
                       storage_name=storage,supplier=supplier)

    return 1;

# 导出指定时间段的入库日志内容
def export_in_storage_log(data):
    storage=data['storage']
    timeSlot=data['datetime']
    fileName=data['fileName']

    table_name="material_in_storage_log"
    if(storage == "全部"):
        if(timeSlot == "本月"):
            result=db.select(table_name, where="DATE_FORMAT(create_time, '%Y%m')=DATE_FORMAT(CURDATE(), '%Y%m')",
                             vars=locals())
        elif(timeSlot == "全部"):
            result=db.select(table_name, vars=locals())
    else:
        if (timeSlot == "本月"):
            result = db.select(table_name,
                               where="DATE_FORMAT(create_time, '%Y%m')=DATE_FORMAT(CURDATE(), '%Y%m') AND storage_name=$storage",
                               vars=locals())
        elif (timeSlot == "全部"):
            result = db.select(table_name, where="storage_name=$storage",vars=locals())

    if(not result):
        return -1;  # 本月无操作

    work_book = xlwt.Workbook(encoding='utf-8')
    sheet = work_book.add_sheet('sheet')
    sheet.write(0, 0, '编号')
    sheet.write(0, 1, '仓库名称')
    sheet.write(0, 2, '类别名称')
    sheet.write(0, 3, '商品名称')
    sheet.write(0, 4, '商品单位')
    sheet.write(0, 5, '入库数量')
    sheet.write(0, 6, '未税单价')
    sheet.write(0, 7, '未税总价')
    sheet.write(0, 8, '发票类型')
    sheet.write(0, 9, '税率')
    sheet.write(0, 10, '含税单价')
    sheet.write(0, 11, '含税总价')
    sheet.write(0, 12, '入库时间')
    logInfo=list(result)
    number=len(logInfo)
    for i in range(1,number+1):
        sheet.write(i, 0, logInfo[i - 1]['id'])
        sheet.write(i, 1, logInfo[i - 1]['storage_name'])
        sheet.write(i, 2, logInfo[i - 1]['category_name'])
        sheet.write(i, 3, logInfo[i - 1]['name'])
        sheet.write(i, 4, logInfo[i - 1]['unit'])
        sheet.write(i, 5, logInfo[i - 1]['count'])
        sheet.write(i, 6, logInfo[i - 1]['price'])
        value = float(Decimal(str(logInfo[i - 1]['price'] * int(logInfo[i - 1]['count']))))
        sheet.write(i, 7, value)
        sheet.write(i, 8, logInfo[i - 1]['invoice_type'])
        value = "%.0f%%" % (logInfo[i - 1]['tax_rate']*100)
        sheet.write(i, 9, value)
        value = float(Decimal(str(logInfo[i - 1]['tax_price']/float(logInfo[i - 1]['count']))))
        sheet.write(i, 10, value)
        sheet.write(i, 11, logInfo[i - 1]['tax_price'])
        value=logInfo[i - 1]['create_time'].strftime("%Y/%m/%d")
        sheet.write(i, 12, value)

    today=datetime.today()
    today=today.strftime("%Y/%m/%d")
    work_book.save(r"C:\Users\ASUS\Desktop\仓库管理系统导出日志\\" + fileName + ".xls")
    #work_book.save(r"C:\Users\ynkmw\Desktop\files\\" + fileName + ".xls")
    return 1

# 导出指定时间段的出库日志内容
def export_out_storage_log(data):
    storage = data['storage']
    timeSlot=data['datetime']
    fileName=data['fileName']

    table_name="material_out_storage_log"

    if (storage == "全部"):
        if (timeSlot == "本月"):
            result = db.select(table_name, where="DATE_FORMAT(create_time, '%Y%m')=DATE_FORMAT(CURDATE(), '%Y%m')",
                               vars=locals())
        elif (timeSlot == "全部"):
            result = db.select(table_name, vars=locals())
    else:
        if (timeSlot == "本月"):
            result = db.select(table_name,
                               where="DATE_FORMAT(create_time, '%Y%m')=DATE_FORMAT(CURDATE(), '%Y%m') AND storage_name=$storage",
                               vars=locals())
        elif (timeSlot == "全部"):
            result = db.select(table_name, where="storage_name=$storage", vars=locals())

    if(not result):
        return -1;  # 本月无操作

    work_book = xlwt.Workbook(encoding='utf-8')
    sheet = work_book.add_sheet('sheet')
    sheet.write(0, 0, '编号')
    sheet.write(0, 1, '申领项目')
    sheet.write(0, 2, '申领仓库')
    sheet.write(0, 3, '实际仓库')
    sheet.write(0, 4, '类别名称')
    sheet.write(0, 5, '商品名称')
    sheet.write(0, 6, '商品单位')
    sheet.write(0, 7, '出库数量')
    sheet.write(0, 8, '出库单价')
    sheet.write(0, 9, '发票类型')
    sheet.write(0, 10, '税率')
    sheet.write(0, 11, '出库总价')
    sheet.write(0, 12, '含税总价')
    sheet.write(0, 13, '出库时间')
    logInfo=list(result)
    number=len(logInfo)
    for i in range(1,number+1):
        sheet.write(i, 0, logInfo[i - 1]['id'])
        sheet.write(i, 1, logInfo[i - 1]['project'])
        sheet.write(i, 2, logInfo[i - 1]['apply_storage'])
        sheet.write(i, 3, logInfo[i - 1]['storage_name'])
        sheet.write(i, 4, logInfo[i - 1]['category_name'])
        sheet.write(i, 5, logInfo[i - 1]['name'])
        sheet.write(i, 6, logInfo[i - 1]['unit'])
        sheet.write(i, 7, logInfo[i - 1]['count'])
        sheet.write(i, 8, logInfo[i - 1]['price'])
        sheet.write(i, 9, logInfo[i - 1]['invoice_type'])
        value = "%.0f%%" % (logInfo[i - 1]['tax_rate'] * 100)
        sheet.write(i, 10, value)
        sheet.write(i, 11, logInfo[i - 1]['total_price'])
        sheet.write(i, 12, logInfo[i - 1]['tax_price'])
        value=logInfo[i - 1]['create_time'].strftime("%Y/%m/%d")
        sheet.write(i, 13, value)

    today=datetime.today()
    today=today.strftime("%Y/%m/%d")
    work_book.save(r"C:\Users\ASUS\Desktop\files\\" + fileName + ".xls")
    #work_book.save(r"C:\Users\ynkmw\Desktop\files\\" + fileName + ".xls")
    return 1

# 导出当前库存信息
def export_storage_info(data):
    storage = data['storage']
    fileName=data['fileName']

    table_name="material_io_storage_info"

    if (storage == "全部"):
        result = db.select(table_name, vars=locals())
    else:
        result = db.select(table_name, where="storage_name=$storage", vars=locals())

    if(not result):
        return -1;  # 本月无操作

    work_book = xlwt.Workbook(encoding='utf-8')
    sheet = work_book.add_sheet('sheet')
    sheet.write(0, 0, '编号')
    sheet.write(0, 1, '实际仓库')
    sheet.write(0, 2, '类别名称')
    sheet.write(0, 3, '商品名称')
    sheet.write(0, 4, '商品单位')
    sheet.write(0, 5, '供应商')
    sheet.write(0, 6, '库存数量')
    sheet.write(0, 7, '未税单价')
    sheet.write(0, 8, '发票类型')
    sheet.write(0, 9, '税率')
    sheet.write(0, 10, '含税总价')
    logInfo=list(result)
    number=len(logInfo)
    for i in range(1,number+1):
        sheet.write(i, 0, logInfo[i - 1]['id'])
        sheet.write(i, 1, logInfo[i - 1]['storage_name'])
        sheet.write(i, 2, logInfo[i - 1]['category_name'])
        sheet.write(i, 3, logInfo[i - 1]['name'])
        sheet.write(i, 4, logInfo[i - 1]['unit'])
        sheet.write(i, 5, logInfo[i - 1]['supplier'])
        sheet.write(i, 6, logInfo[i - 1]['count'])
        sheet.write(i, 7, logInfo[i - 1]['price'])
        sheet.write(i, 8, logInfo[i - 1]['invoice_type'])
        value = "%.0f%%" % (logInfo[i - 1]['tax_rate'] * 100)
        sheet.write(i, 9, value)
        sheet.write(i, 10, logInfo[i - 1]['tax_price'])

    today=datetime.today()
    today=today.strftime("%Y/%m/%d")
    work_book.save(r"C:\Users\ASUS\Desktop\files\\" + fileName + ".xls")
    #work_book.save(r"C:\Users\ynkmw\Desktop\files\\" + fileName + ".xls")
    return 1

# 筛选指定时间段的入库日志内容
def filter_in_storage_log(data):
    storage=data['storage']
    timeSlot=data['datetime']

    table_name = "material_in_storage_log"
    #未指定仓库
    if (storage == "全部"):
        if (timeSlot == "本月"):
            result = list(db.select(table_name, where="DATE_FORMAT(create_time, '%Y%m')=DATE_FORMAT(CURDATE(), '%Y%m')",
                               order="id DESC",vars=locals()))
        elif (timeSlot == "全部"):
            result = list(db.select(table_name, order="id DESC",vars=locals()))
    #指定仓库
    else:
        if (timeSlot == "本月"):
            result = list(db.select(table_name,
                               where="DATE_FORMAT(create_time, '%Y%m')=DATE_FORMAT(CURDATE(), '%Y%m') AND storage_name=$storage",
                               order="id DESC",vars=locals()))
        elif (timeSlot == "全部"):
            result = list(db.select(table_name, where="storage_name=$storage", order="id DESC",vars=locals()))

    return result


# 筛选指定时间段的出库日志内容
def filter_out_storage_log(data):
    storage = data['storage']
    timeSlot = data['datetime']

    table_name = "material_out_storage_log"
    #未指定仓库
    if (storage == "全部"):
        if (timeSlot == "本月"):
            result = list(db.select(table_name, where="DATE_FORMAT(create_time, '%Y%m')=DATE_FORMAT(CURDATE(), '%Y%m')",
                               order="id DESC",vars=locals()))
        elif (timeSlot == "全部"):
            result = list(db.select(table_name, order="id DESC",vars=locals()))
    #指定仓库
    else:
        if (timeSlot == "本月"):
            result = list(db.select(table_name,
                               where="DATE_FORMAT(create_time, '%Y%m')=DATE_FORMAT(CURDATE(), '%Y%m') AND storage_name=$storage",
                                    order="id DESC", vars=locals()))
        elif (timeSlot == "全部"):
            result = list(db.select(table_name, order="id DESC",where="storage_name=$storage", vars=locals()))

    return result

# 筛选仓库中符合还货要求的商品信息
def filter_stock_borrow_info(data):
    # 欠货的仓库
    borrow_storage = data['storage']
    # 欠货数量
    count = data['count']
    # 商品类型
    category = data['category']
    # 商品名称
    material = data['material']

    table_name = "material_io_storage_info"
    result = list(db.select(table_name,
                       where="category_name=$category AND name=$material AND storage_name=$borrow_storage AND count>=$count",
                       vars=locals()))
    return result

# 还货功能
def material_returned_info(data):

    # 借货出库的出库记录id
    borrow_id = data['borrow_id']

    # 用来还货的库存记录id
    stock_id = data['stock_id']

    table_name="material_out_storage_log"
    result = list(db.select(table_name,where="id=$borrow_id",vars=locals()))

    borrow_info = result[0]

    borrow_count = borrow_info['count']           # 借货数量
    borrow_material = borrow_info['name']         # 借货商品
    borrow_supplier = borrow_info['supplier']     # 借货供应商
    borrow_batch = borrow_info['material_batch']  # 借货批次
    borrow_tax_price = borrow_info['tax_price']       # 借货的含税总金额


    '''     1.将调货的那笔库存，加回去      '''
    '''     1.1 通过批次（时间戳）来找到当时借用的哪一笔材料      '''
    table_name="material_io_storage_info"
    result = list(db.select(table_name,where="create_time=$borrow_batch AND name=$borrow_material",vars=locals()))
    if(len(result)==1):
        new_count = result[0]['count']+borrow_count
        new_tax_price = result[0]['tax_price'] + borrow_tax_price
        #new_tax_price = float(Decimal(str(result[0]['tax_price'] / result[0]['count'] * new_count)))
        db.update(table_name,where="create_time=$borrow_batch AND name=$borrow_material",
                  vars=locals(),count=new_count,tax_price=new_tax_price)
    else:
        return -1   # 库存更新出错

    '''     1.2 存入出库日志表(将借货出库的那笔日志只更改数量再次存储)     '''
    table_name = "material_out_storage_log"
    category_id = borrow_info['category_id']
    category_name = borrow_info['category_name']
    material_id = borrow_info['material_id']
    material_name = borrow_info['name']
    unit = borrow_info['unit']
    outCount = -borrow_count
    price = borrow_info['price']
    invoice = borrow_info['invoice_type']
    tax = borrow_info['tax_rate']
    total_price = -borrow_info['total_price']
    tax_price = -borrow_info['tax_price']
    project=borrow_info['project']
    supplier=borrow_info['supplier']
    storage = borrow_info['storage_name']
    borrow_storage = borrow_info['apply_storage']
    batch = borrow_info['material_batch']
    statusValue="RETURN-"+str(borrow_id)     #数据库记录还的哪一笔
    status = db.insert(table_name,
                       category_id=category_id, category_name=category_name, material_id=material_id,
                       name=material_name, unit=unit, count=outCount, price=price, invoice_type=invoice, tax_rate=tax,
                       total_price=total_price, tax_price=tax_price, project=project, storage_name=storage,
                       supplier=supplier, apply_storage=borrow_storage, material_batch=batch,status=statusValue
                       )


    '''     2.更改还货的那条库存信息        '''

    table_name = "material_io_storage_info"
    # 用还还货的那条库存信息
    stock_material = list(db.select(table_name, where="id=$stock_id", vars=locals()))

    '''     2.1更新该条入库信息状态，不允许删除        '''
    # 该条商品入库信息的状态（入库时间一致且商品名称一致，库存数量与入库数量一致）
    current_status = list(db.select("material_in_storage_log",
                                    where="create_time=$stock_material[0]['create_time'] AND name=$stock_material[0]['name'] AND count=$stock_material[0]['count']",
                                    vars=locals()))
    # 0未出过库 1已出库 -1已手动删除
    if len(current_status) == 1 and current_status[0]['status'] == 0:
        status = db.update("material_in_storage_log",
                  where="create_time=$stock_material[0]['create_time'] AND name=$stock_material[0]['name'] AND count=$stock_material[0]['count']",
                  vars=locals(), status=1)  # 更新状态标志位
        if (not status ):
            return -4

    '''     2.2还货的这笔库存减少        '''

    # 计算该条商品信息的扣减完的库存
    current_count = stock_material[0]['count']
    new_count = int(current_count) - int(borrow_count)  # 当前数量减去出库数量

    # 计算该条商品的最新入库含税总价
    current_tax_price = stock_material[0]['tax_price']
    new_tax_price = current_tax_price / current_count * new_count  # 获得原先含税单价并乘上现有个数

    status = db.update(table_name, where="id=$stock_id", vars=locals(), count=new_count, tax_price=new_tax_price)
    if (not status):
        return -2     # 更新库存信息


    '''     2.3更新出库日志       '''
    if(not update_material_out_log(stock_material,borrow_count,project,statusValue)):
        return -3      #更新出库日志出错

    '''     3更新该日志消息为已还     '''
    status = db.update("material_out_storage_log",where="id=$borrow_id",vars=locals(),status="RETURNED")
    if (not status):
        return 0
    return 1



def update_material_out_log(stock_material,borrow_count,project,statusValue):

    table_name = "material_out_storage_log"
    stock_material_info = stock_material[0]

    batch = stock_material_info['create_time']
    category_id = stock_material_info['category_id']
    category_name = stock_material_info['category_name']
    material_id = stock_material_info['material_id']
    material_name = stock_material_info['name']
    unit = stock_material_info['unit']
    outCount = borrow_count
    price = stock_material_info['price']
    tax = stock_material_info['tax_rate']
    invoice = stock_material_info['invoice_type']
    storage = stock_material_info['storage_name']
    supplier = stock_material_info['supplier']

    value = float(Decimal(str(1 + tax)) * Decimal(str(price)))  # 单个含税价
    tax_price = float(Decimal(str(value * int(outCount))))  # 含税总价

    # 专票商品出库时的金额采用入库的未税单价金额
    if (stock_material_info['invoice_type'] == "专票"):
        total_price = float(Decimal(str(price * int(outCount))))
    else:
        price = value
        total_price = tax_price

    status = db.insert(table_name,
                       category_id=category_id, category_name=category_name, material_id=material_id,
                       name=material_name, unit=unit, count=outCount, price=price, invoice_type=invoice, tax_rate=tax,
                       total_price=total_price, tax_price=tax_price, project=project, storage_name=storage,
                       supplier=supplier, apply_storage=storage, material_batch=batch,status=statusValue
                       )
    if( not status):
        return 0
    else:
        return 1
